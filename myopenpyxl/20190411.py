# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet('sheet1')
ws2 = wb.create_sheet('sheet2')
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

target = wb.copy_worksheet(ws)
wb.save('./data/1.xlsx')
wb.close()